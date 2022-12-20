import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

wb = Workbook()
ws = wb.active

meisai = wb.create_sheet('明細データ')

df = pd.read_csv('./data/matome.csv', encoding='shift-jis',
                 usecols=['発生日', '発生時刻', '射出最前進位置[mm]', 'V-P切換圧[MPa]', '射出最前進圧[MPa]', '射出ピーク圧[MPa]'])


today_df = df.query('発生日 >= "2022/12/20"')


today_df['発生日'] = pd.to_datetime(
    today_df['発生日'] + ' ' + today_df['発生時刻'])

select_df = today_df[['発生日', '射出最前進位置[mm]',
                     'V-P切換圧[MPa]', '射出最前進圧[MPa]', '射出ピーク圧[MPa]']]

select_df.columns = ['発生日時', '射出最前進位置[mm]', 'V-P切換圧[MPa]',
                     '射出最前進位置[mm]しきい値', 'V-P切換圧[MPa]しきい値']

select_df['射出最前進位置[mm]しきい値'] = 1.89
select_df['V-P切換圧[MPa]しきい値'] = 47.3

select_df1 = select_df[['発生日時', '射出最前進位置[mm]',
                        '射出最前進位置[mm]しきい値', '発生日時', 'V-P切換圧[MPa]', 'V-P切換圧[MPa]しきい値']]


for row in dataframe_to_rows(select_df1, index=None, header=True):
    meisai.append(row)


# 射出最前進位置作成
line = LineChart()
line.style = 26
line.height = 17
line.width = 50
data = Reference(meisai, min_col=2, max_col=3,
                 min_row=1, max_row=meisai.max_row)
labels = Reference(meisai, min_col=1, min_row=2, max_row=meisai.max_row)

line.add_data(data, titles_from_data=True)
line.set_categories(labels)

line.x_axis.title = '射出最前進位置のしきい値は「1.89mm」　これ以上はショートリスク有り！'


ws.add_chart(line, 'A1')

# VP切替圧のグラフ作成

line2 = LineChart()
line2.style = 26
line2.height = 17
line2.width = 50
data2 = Reference(meisai, min_col=5, max_col=6,
                  min_row=1, max_row=meisai.max_row)
labels2 = Reference(meisai, min_col=4, min_row=2, max_row=meisai.max_row)

line2.add_data(data2, titles_from_data=True)
line2.set_categories(labels2)

line2.x_axis.title = 'VP切替圧のしきい値は「47.3Mpa」　これ以下はショートリスク有り！'


ws.add_chart(line2, 'A40')


wb.save('H:\マイドライブ\成形機モニターデータ\横浜モニターデータ\横浜モニターデータ.xlsx')
